import sys
import os

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from lncfit.screen_data import (
    ScreenRecord, load_targets, load_annotations, load_screen, to_dataframe,
    save_jsonl, load_jsonl, SCHEMA_VERSION,
)


_FC_HEADERS = [
    "Day 7 Replicate 1 (Fold-change)",
    "Day 7 Replicate 2 (Fold-change)",
    "Day 14 Replicate 1 (Fold-change)",
    "Day 14 Replicate 2 (Fold-change)",
]

_S2_SHEETS = ["S2A", "S2B", "S2C", "S2D", "S2E"]


def _make_mmc2(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_a = wb.create_sheet("S1A")
    ws_a.append(["Table S1A title"])
    ws_a.append([])
    ws_a.append([
        "lncRNA", "ENSEMBL_ID", "Chr", "Strand",
        "Closest protein-coding gene symbol",
        "Distance to closest protein-coding gene",
    ])
    ws_a.append(["Hum_GENE1", "ENSG001", "1", "+", "GENE_A", 18435])
    ws_a.append(["Hum_GENE2", "ENSG002", "X", "-", "GENE_B", None])

    ws_b = wb.create_sheet("S1B")
    ws_b.append(["ID", "target", "sequence"])
    ws_b.append(["gL_000001", "Hum_GENE1", "AAACCCGGGTTT"])
    ws_b.append(["gL_000002", "Hum_GENE2", "TTTAGCGCGCGC"])

    path = tmp_path / "mmc2.xlsx"
    wb.save(path)
    return path


def _make_mmc3(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet in _S2_SHEETS:
        ws = wb.create_sheet(sheet)
        ws.append(["ID"] + _FC_HEADERS)
        ws.append(["gL_000001", 0.5, 0.6, -0.3, -0.4])
        ws.append(["gL_000002", 1.2, 1.1, 0.8, 0.9])
    path = tmp_path / "mmc3.xlsx"
    wb.save(path)
    return path


def test_load_targets(tmp_path):
    targets = load_targets(_make_mmc2(tmp_path))
    assert targets["gL_000001"] == ("Hum_GENE1", "AAACCCGGGTTT")
    assert targets["gL_000002"] == ("Hum_GENE2", "TTTAGCGCGCGC")


def test_load_annotations(tmp_path):
    annots = load_annotations(_make_mmc2(tmp_path))
    assert annots["Hum_GENE1"] == ("1", "+", "GENE_A", 18435)
    assert annots["Hum_GENE2"] == ("X", "-", "GENE_B", None)


def test_load_annotations_blank_distance_is_none(tmp_path):
    annots = load_annotations(_make_mmc2(tmp_path))
    assert annots["Hum_GENE2"][3] is None


def test_melt_produces_4_rows_per_guide_per_cell_line(tmp_path):
    mmc2 = _make_mmc2(tmp_path)
    targets = load_targets(mmc2)
    records = load_screen(_make_mmc3(tmp_path), targets)
    subset = [r for r in records if r.guide_id == "gL_000001" and r.cell_line == "HAP1"]
    assert len(subset) == 4


def test_total_row_count(tmp_path):
    mmc2 = _make_mmc2(tmp_path)
    targets = load_targets(mmc2)
    records = load_screen(_make_mmc3(tmp_path), targets)
    # 2 guides × 5 cell lines × 4 FC columns = 40
    assert len(records) == 40


def test_day_and_replicate_parsed(tmp_path):
    mmc2 = _make_mmc2(tmp_path)
    targets = load_targets(mmc2)
    records = load_screen(_make_mmc3(tmp_path), targets)
    assert {r.day for r in records} == {7, 14}
    assert {r.replicate for r in records} == {1, 2}


def test_negative_fold_changes_preserved(tmp_path):
    mmc2 = _make_mmc2(tmp_path)
    targets = load_targets(mmc2)
    records = load_screen(_make_mmc3(tmp_path), targets)
    assert any(r.fold_change < 0 for r in records)


def test_s1_s2_join(tmp_path):
    mmc2 = _make_mmc2(tmp_path)
    targets = load_targets(mmc2)
    records = load_screen(_make_mmc3(tmp_path), targets)
    r = next(r for r in records if r.guide_id == "gL_000001")
    assert r.target == "Hum_GENE1"
    assert r.target_sequence == "AAACCCGGGTTT"


def test_annotations_enriched_in_records(tmp_path):
    mmc2 = _make_mmc2(tmp_path)
    targets = load_targets(mmc2)
    annots = load_annotations(mmc2)
    records = load_screen(_make_mmc3(tmp_path), targets, annotations=annots)
    r = next(r for r in records if r.guide_id == "gL_000001")
    assert r.chrom == "1"
    assert r.strand == "+"
    assert r.closest_pc_gene == "GENE_A"
    assert r.distance_to_closest_pc_gene == 18435


def test_missing_annotation_falls_back_to_defaults(tmp_path):
    mmc2 = _make_mmc2(tmp_path)
    targets = load_targets(mmc2)
    # Pass empty annotations dict — no match for any target
    records = load_screen(_make_mmc3(tmp_path), targets, annotations={})
    r = records[0]
    assert r.chrom == ""
    assert r.strand == ""
    assert r.closest_pc_gene == ""
    assert r.distance_to_closest_pc_gene is None


def test_to_dataframe_schema(tmp_path):
    mmc2 = _make_mmc2(tmp_path)
    targets = load_targets(mmc2)
    annots = load_annotations(mmc2)
    records = load_screen(_make_mmc3(tmp_path), targets, annotations=annots)
    df = to_dataframe(records)
    assert set(df.columns) == {
        "guide_id", "target", "target_sequence",
        "cell_line", "day", "replicate", "fold_change",
        "chrom", "strand", "closest_pc_gene", "distance_to_closest_pc_gene",
    }
    assert len(df) == 40


def test_save_load_jsonl_round_trip(tmp_path):
    mmc2 = _make_mmc2(tmp_path)
    targets = load_targets(mmc2)
    annots = load_annotations(mmc2)
    records = load_screen(_make_mmc3(tmp_path), targets, annotations=annots)
    path = tmp_path / "records.jsonl.gz"
    save_jsonl(records, path)
    loaded = load_jsonl(path)
    assert len(loaded) == len(records)
    assert loaded[0] == records[0]


def test_jsonl_stamped_with_schema_version(tmp_path):
    import gzip
    import json
    record = ScreenRecord("g1", "T1", "ACGT", "HAP1", 7, 1, -1.0)
    path = tmp_path / "records.jsonl.gz"
    save_jsonl([record], path)
    with gzip.open(path, "rt", encoding="utf-8") as f:
        line = json.loads(f.read().strip())
    assert line["_v"] == SCHEMA_VERSION


def test_from_dict_ignores_unknown_keys():
    d = {
        "guide_id": "g1", "target": "T1", "target_sequence": "ACGT",
        "cell_line": "HAP1", "day": 7, "replicate": 1, "fold_change": -1.0,
        "future_field": "ignored",
    }
    r = ScreenRecord.from_dict(d)
    assert r.guide_id == "g1"
    assert r.chrom == ""  # default applied


def test_from_dict_missing_optional_fields_use_defaults():
    d = {
        "guide_id": "g1", "target": "T1", "target_sequence": "ACGT",
        "cell_line": "HAP1", "day": 7, "replicate": 1, "fold_change": -1.0,
    }
    r = ScreenRecord.from_dict(d)
    assert r.chrom == ""
    assert r.strand == ""
    assert r.closest_pc_gene == ""
    assert r.distance_to_closest_pc_gene is None
