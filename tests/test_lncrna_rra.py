import sys
import os
import gzip
import json

import openpyxl
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from lncfit.screen_data import (
    LncRnaRecord, load_target_groups, load_rra, load_annotations,
    load_jsonl, save_jsonl, SCHEMA_VERSION,
)


_RRA_SHEETS = ["S2F", "S2G", "S2H", "S2I", "S2J"]
_RRA_HEADER = [
    "Gene", "Day 7 - P value", "Day 7 - Fold-change (log2)",
    "Day 14 - P value", "Day 14 - Fold-change (log2)",
]


def _make_mmc2(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_a = wb.create_sheet("S1A")
    ws_a.append(["Table S1A title"])
    ws_a.append([])
    ws_a.append([
        "lncRNA", "ENSEMBL_ID", "Chr", "Strand",
        "Closest protein-coding gene symbol",
        "Distance to closest protein-coding gene",
    ])
    ws_a.append(["Hum_XLOC_000001", "ENSG001", "1", "+", "GENE_A", 18435])
    ws_a.append(["Hum_XLOC_000002", "ENSG002", "X", "-", "GENE_B", None])

    ws_b = wb.create_sheet("S1B")
    ws_b.append(["Table S1B title"])
    ws_b.append([])
    ws_b.append(["ID", "Target", "Sequence (5' - 3')", "Target group"])
    ws_b.append(["gL_000001", "Hum_XLOC_000001", "AAACCCGGGTTT", "long non-coding RNA"])
    ws_b.append(["gL_000002", "Hum_XLOC_000001", "GGGTTTAAACCC", "long non-coding RNA"])
    ws_b.append(["gL_000003", "Hum_XLOC_000002", "TTTAGCGCGCGC", "long non-coding RNA"])
    ws_b.append(["gL_000004", "TP53", "CATGCATGCATG", "protein-coding gene"])
    ws_b.append(["gL_000005", "NTC_001", "AAAAAAAAAAAA", "non-targeting"])

    path = tmp_path / "mmc2.xlsx"
    wb.save(path)
    return path


def _make_mmc3(tmp_path, rows_by_sheet=None):
    """rows_by_sheet: {sheet_name: [[gene, d7_p, d7_fc, d14_p, d14_fc], ...]}. Defaults to a
    shared row set (one lncRNA hit, one lncRNA non-hit, one protein-coding gene) per sheet."""
    default_rows = [
        ["Hum_XLOC_000001", 0.01, -1.2, 0.02, -0.9],   # significant + negative -> hit
        ["Hum_XLOC_000002", 0.9, -0.5, 0.8, -0.4],     # not significant -> non-hit
        ["TP53", 0.01, -2.0, 0.01, -2.0],              # protein-coding gene -> excluded
    ]
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet in _RRA_SHEETS:
        ws = wb.create_sheet(sheet)
        ws.append(_RRA_HEADER)
        if rows_by_sheet is None:
            rows = default_rows
        else:
            rows = rows_by_sheet.get(sheet, [])
        for row in rows:
            ws.append(row)
    path = tmp_path / "mmc3.xlsx"
    wb.save(path)
    return path


def test_load_target_groups(tmp_path):
    groups = load_target_groups(_make_mmc2(tmp_path))
    assert groups["Hum_XLOC_000001"] == "long non-coding RNA"
    assert groups["TP53"] == "protein-coding gene"
    assert groups["NTC_001"] == "non-targeting"


def test_load_rra_filters_to_lncrna_only(tmp_path):
    mmc2 = _make_mmc2(tmp_path)
    groups = load_target_groups(mmc2)
    records = load_rra(_make_mmc3(tmp_path), day=14, target_groups=groups)
    # 2 lncRNAs x 5 cell lines = 10; TP53 rows are dropped entirely, not relabeled
    assert len(records) == 10
    assert {r.target for r in records} == {"Hum_XLOC_000001", "Hum_XLOC_000002"}


# The 3 branches of the compound hit condition (p<0.05 AND log2FC<0) are each
# independently tested -- flipping either half of the AND is a classic bug.
def test_load_rra_label_significant_and_negative_is_hit(tmp_path):
    mmc2 = _make_mmc2(tmp_path)
    groups = load_target_groups(mmc2)
    records = load_rra(_make_mmc3(tmp_path), day=14, target_groups=groups)
    r = next(r for r in records if r.target == "Hum_XLOC_000001" and r.cell_line == "HAP1")
    assert r.rra_pvalue == pytest.approx(0.02)
    assert r.fold_change == pytest.approx(-0.9)
    assert r.label == 1


def test_load_rra_label_not_significant_is_non_hit(tmp_path):
    mmc2 = _make_mmc2(tmp_path)
    groups = load_target_groups(mmc2)
    records = load_rra(_make_mmc3(tmp_path), day=14, target_groups=groups)
    r = next(r for r in records if r.target == "Hum_XLOC_000002" and r.cell_line == "HAP1")
    assert r.label == 0


def test_load_rra_significant_but_positive_fc_is_non_hit(tmp_path):
    mmc2 = _make_mmc2(tmp_path)
    groups = load_target_groups(mmc2)
    rows = {"S2F": [["Hum_XLOC_000001", 0.01, 0.8, 0.01, 0.8]]}
    records = load_rra(_make_mmc3(tmp_path, rows), day=14, target_groups=groups)
    assert len(records) == 1
    assert records[0].label == 0


def test_load_rra_day_selection_reads_correct_columns(tmp_path):
    mmc2 = _make_mmc2(tmp_path)
    groups = load_target_groups(mmc2)
    mmc3 = _make_mmc3(tmp_path)
    day7 = load_rra(mmc3, day=7, target_groups=groups)
    day14 = load_rra(mmc3, day=14, target_groups=groups)
    r7 = next(r for r in day7 if r.target == "Hum_XLOC_000001" and r.cell_line == "HAP1")
    r14 = next(r for r in day14 if r.target == "Hum_XLOC_000001" and r.cell_line == "HAP1")
    assert r7.day == 7 and r7.rra_pvalue == pytest.approx(0.01) and r7.fold_change == pytest.approx(-1.2)
    assert r14.day == 14 and r14.rra_pvalue == pytest.approx(0.02) and r14.fold_change == pytest.approx(-0.9)


def test_load_rra_annotation_join_and_missing_defaults(tmp_path):
    mmc2 = _make_mmc2(tmp_path)
    groups = load_target_groups(mmc2)
    annots = load_annotations(mmc2)
    records = load_rra(_make_mmc3(tmp_path), day=14, target_groups=groups, annotations=annots)
    r = next(r for r in records if r.target == "Hum_XLOC_000001" and r.cell_line == "HAP1")
    assert r.chrom == "1" and r.strand == "+" and r.closest_pc_gene == "GENE_A"
    assert r.distance_to_closest_pc_gene == 18435

    no_annot_records = load_rra(_make_mmc3(tmp_path), day=14, target_groups=groups, annotations={})
    r2 = no_annot_records[0]
    assert r2.chrom == "" and r2.strand == "" and r2.closest_pc_gene == ""
    assert r2.distance_to_closest_pc_gene is None


def test_load_rra_skips_missing_pvalue_or_fc(tmp_path):
    mmc2 = _make_mmc2(tmp_path)
    groups = load_target_groups(mmc2)
    rows = {"S2F": [["Hum_XLOC_000001", None, None, 0.01, -1.0]]}
    records = load_rra(_make_mmc3(tmp_path, rows), day=7, target_groups=groups)
    assert records == []


def test_lncrna_jsonl_round_trip_and_schema_version(tmp_path):
    mmc2 = _make_mmc2(tmp_path)
    groups = load_target_groups(mmc2)
    records = load_rra(_make_mmc3(tmp_path), day=14, target_groups=groups)
    path = tmp_path / "lncrna.jsonl.gz"
    save_jsonl(records, path)
    loaded = load_jsonl(path, record_cls=LncRnaRecord)
    assert len(loaded) == len(records)
    assert loaded[0] == records[0]
    with gzip.open(path, "rt", encoding="utf-8") as f:
        assert json.loads(f.readline())["_v"] == SCHEMA_VERSION


def test_lncrna_from_dict_unknown_and_missing_fields():
    base = {
        "target": "Hum_XLOC_1", "cell_line": "HAP1", "day": 14,
        "rra_pvalue": 0.01, "fold_change": -1.0, "label": 1,
    }
    r = LncRnaRecord.from_dict({**base, "future_field": "ignored"})
    assert r.target == "Hum_XLOC_1" and r.chrom == ""  # unknown key ignored, default applied

    r = LncRnaRecord.from_dict(base)  # all optional fields omitted entirely
    assert r.chrom == "" and r.strand == "" and r.closest_pc_gene == ""
    assert r.distance_to_closest_pc_gene is None
